from django.contrib import admin
from django.db.models import Sum
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.text import slugify

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

import openpyxl
import csv
from io import BytesIO

from .models import Vaccin


# ------------------------- Filtres Personnalisés ------------------------- #

class MoisFilter(admin.SimpleListFilter):
    title = 'Mois'
    parameter_name = 'mois'

    def lookups(self, request, model_admin):
        mois = model_admin.model.objects.values_list('mois', flat=True).distinct()
        return [(m, m) for m in mois if m]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(mois=self.value())
        return queryset


class AnneeFilter(admin.SimpleListFilter):
    title = 'Année'
    parameter_name = 'annee'

    def lookups(self, request, model_admin):
        annees = model_admin.model.objects.values_list('annee', flat=True).distinct()
        return [(a, a) for a in annees if a]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(annee=self.value())
        return queryset


class ActeurFilterBase(admin.SimpleListFilter):
    def lookups(self, request, model_admin):
        values = model_admin.model.objects.values_list(self.parameter_name, flat=True).distinct()
        return [(v, v) for v in values if v]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(**{self.parameter_name: self.value()})
        return queryset


class Acteur1Filter(ActeurFilterBase):
    title = 'Acteur 1'
    parameter_name = 'nom_acteur1'


class Acteur2Filter(ActeurFilterBase):
    title = 'Acteur 2'
    parameter_name = 'nom_acteur2'


class Acteur3Filter(ActeurFilterBase):
    title = 'Acteur 3'
    parameter_name = 'nom_acteur3'


# ------------------------- Base Admin Générique ------------------------- #

class BaseAdmin(admin.ModelAdmin):
    actions = ['export_as_excel', 'export_as_pdf', 'export_as_csv']

    def get_export_fields(self, request):
        return []

    def get_export_headers(self, request):
        return []

    def get_export_filename_prefix(self, request):
        return self.model._meta.model_name

    def get_export_pdf_title(self, request):
        return self.model._meta.verbose_name_plural.capitalize()

    def export_as_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        fields = self.get_export_fields(request)
        headers = self.get_export_headers(request) or fields
        ws.append(headers)

        for obj in queryset:
            ws.append([getattr(obj, field) for field in fields])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{slugify(self.get_export_filename_prefix(request))}.xlsx"'
        wb.save(response)
        return response

    def export_as_pdf(self, request, queryset):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 50

        fields = self.get_export_fields(request)
        title = self.get_export_pdf_title(request)

        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, title)
        y -= 30
        p.setFont("Helvetica", 10)

        for obj in queryset:
            if y < 50:
                p.showPage()
                y = height - 50
                p.setFont("Helvetica", 10)
            p.drawString(50, y, " | ".join(str(getattr(obj, f)) for f in fields))
            y -= 20

        p.showPage()
        p.save()
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{slugify(self.get_export_filename_prefix(request))}.pdf"'
        return response

    def export_as_csv(self, request, queryset):
        fields = self.get_export_fields(request)
        headers = self.get_export_headers(request) or fields

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{slugify(self.get_export_filename_prefix(request))}.csv"'

        writer = csv.writer(response, delimiter=';')
        writer.writerow(headers)

        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in fields])

        return response


# ------------------------- Admin du modèle Vaccin ------------------------- #

@admin.register(Vaccin)
class VaccinAdmin(BaseAdmin):
    list_display = (
        'libelle', 'mois', 'annee', 'montantTT', 'maison',
        'nom_acteur1', 'acteur1_part',
        'nom_acteur2', 'acteur2_part',
        'nom_acteur3', 'acteur3_part',
    )
    readonly_fields = ('maison', 'acteur1_part', 'acteur2_part', 'acteur3_part')
    list_filter = [MoisFilter, AnneeFilter, Acteur1Filter, Acteur2Filter, Acteur3Filter]
    ordering = ['-annee', '-mois']

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(request, extra_context)
        if hasattr(response, 'context_data') and 'cl' in response.context_data:
            cl = response.context_data['cl']
            qs = cl.queryset

            total = sum((obj.montantTT or 0) for obj in qs)
            maison = sum((obj.maison or 0) for obj in qs)
            acteur1 = sum((obj.acteur1_part or 0) for obj in qs)
            acteur2 = sum((obj.acteur2_part or 0) for obj in qs)
            acteur3 = sum((obj.acteur3_part or 0) for obj in qs)

            self.message_user(
                request,
                format_html(
                    "<b>Totaux filtrés :</b> "
                    "Montant Total : <b>{}</b> | "
                    "Maison : <b>{}</b> | "
                    "Acteur 1 : <b>{}</b> | "
                    "Acteur 2 : <b>{}</b> | "
                    "Acteur 3 : <b>{}</b>",
                    round(total, 2),
                    round(maison, 2),
                    round(acteur1, 2),
                    round(acteur2, 2),
                    round(acteur3, 2),
                )
            )
        return response

    def get_export_fields(self, request):
        return [
            'libelle', 'mois', 'annee', 'montantTT', 'maison',
            'nom_acteur1', 'acteur1_part',
            'nom_acteur2', 'acteur2_part',
            'nom_acteur3', 'acteur3_part',
        ]

    def get_export_headers(self, request):
        return [
            'Libellé', 'Mois', 'Année', 'Montant Total', 'Maison',
            'Nom Acteur 1', 'Part Acteur 1',
            'Nom Acteur 2', 'Part Acteur 2',
            'Nom Acteur 3', 'Part Acteur 3',
        ]

    def get_export_filename_prefix(self, request):
        return "vaccins"

    def get_export_pdf_title(self, request):
        return "Rapport Vaccins"

    def export_as_excel(self, request, queryset):
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        fields = self.get_export_fields(request)
        headers = self.get_export_headers(request)
        ws.append(headers)

        for obj in queryset:
            ws.append([getattr(obj, field) for field in fields])

        # Ligne des totaux
        total_row = ["Total", "", ""]  # pour libellé, mois, année
        total_row += [
            sum((obj.montantTT or 0) for obj in queryset),
            sum((obj.maison or 0) for obj in queryset),
            "", sum((obj.acteur1_part or 0) for obj in queryset),
            "", sum((obj.acteur2_part or 0) for obj in queryset),
            "", sum((obj.acteur3_part or 0) for obj in queryset),
        ]
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{slugify(self.get_export_filename_prefix(request))}.xlsx"'
        wb.save(response)
        return response

    def export_as_pdf(self, request, queryset):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 50

        fields = self.get_export_fields(request)
        title = self.get_export_pdf_title(request)

        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, title)
        y -= 30
        p.setFont("Helvetica", 10)

        for obj in queryset:
            if y < 50:
                p.showPage()
                y = height - 50
                p.setFont("Helvetica", 10)
            p.drawString(50, y, " | ".join(str(getattr(obj, f)) for f in fields))
            y -= 20

        # Ligne Totaux
        if y < 70:
            p.showPage()
            y = height - 50
        p.setFont("Helvetica-Bold", 10)
        total_line = [
            "TOTAL", "", "",
            round(sum((obj.montantTT or 0) for obj in queryset), 2),
            round(sum((obj.maison or 0) for obj in queryset), 2),
            "", round(sum((obj.acteur1_part or 0) for obj in queryset), 2),
            "", round(sum((obj.acteur2_part or 0) for obj in queryset), 2),
            "", round(sum((obj.acteur3_part or 0) for obj in queryset), 2),
        ]
        p.drawString(50, y, " | ".join(str(val) for val in total_line))

        p.showPage()
        p.save()
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{slugify(self.get_export_filename_prefix(request))}.pdf"'
        return response

    def export_as_csv(self, request, queryset):
        import csv
        from django.utils.text import slugify

        fields = self.get_export_fields(request)
        headers = self.get_export_headers(request)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{slugify(self.get_export_filename_prefix(request))}.csv"'

        writer = csv.writer(response, delimiter=';')
        writer.writerow(headers)

        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in fields])

        # Totaux CSV
        writer.writerow([
            'TOTAL', '', '',
            round(sum((obj.montantTT or 0) for obj in queryset), 2),
            round(sum((obj.maison or 0) for obj in queryset), 2),
            '', round(sum((obj.acteur1_part or 0) for obj in queryset), 2),
            '', round(sum((obj.acteur2_part or 0) for obj in queryset), 2),
            '', round(sum((obj.acteur3_part or 0) for obj in queryset), 2),
        ])

        return response
